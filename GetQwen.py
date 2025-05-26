import os
import time
import base64
import threading
import multiprocessing
from pathlib import Path
from queue import Queue
from openai import OpenAI
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
from tqdm import tqdm

# ====== OpenAI 阿里云千问客户端 ======
client = OpenAI(
    api_key="sk-038c3b8e36ce43e4a94156257a766bc4",
    base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
)

# ====== 配置路径与参数 ======
IMAGE_FOLDER = "dataset/images"
OUTPUT_EXCEL = "dataset/image_descriptions.xlsx"
VALID_EXTENSIONS = {".jpg", ".png", ".jpeg"}
MAX_THREADS = min(20, multiprocessing.cpu_count() * 2)
MAX_RETRIES = 3  # 最大重试次数

# ====== 初始化 Excel 工作簿 ======
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "图片描述"
ws.append(["图片缩略图", "图片路径", "图片描述"])

lock = threading.Lock()
task_queue = Queue()

# ====== 统计数据 ======
total_images = 0
processed_count = 0
success_count = 0
failure_count = 0
start_time = time.time()

def describe_image(image_path, client):
    for attempt in range(MAX_RETRIES):
        try:
            with open(image_path, "rb") as image_file:
                base64_image = base64.b64encode(image_file.read()).decode("utf-8")

            response = client.chat.completions.create(
                model="qwen2.5-vl-72b-instruct",
                messages=[
                    {"role": "user", "content": [
                        {"type": "text", "text": "你是一位机械制造与装配优化领域的专家。我将提供一张图片，展示一个用于SMA（表面贴装组件）装配的盒子。该盒子共有20个空位，排布为4行5列，依次编号为 (1,1) 到 (4,5)，其中有2个空位已完成安装。请你根据图片准确识别所有空位的位置与编号，特别注意不要识别错已安装的位置，务必确保它们从后续安装顺序中排除。在此基础上，请以最小化工具移动距离为目标，为剩余18个空位规划一个最优的安装顺序。最终，请按照以下格式输出每个空位的安装顺序及其在图像中的坐标：格式示例：(1,3) ，(1,4) ..(4,1) 。请确保输出顺序合理，并尽可能减少装配过程中的路径长度。"},
                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}}
                    ]}
                ]
            )
            return response.choices[0].message.content
        except Exception as e:
            print(f"[❌ ERROR] 第 {attempt + 1} 次尝试处理 {image_path} 时出错: {e}")
            time.sleep(1)
    return None

def worker(progress_bar):
    global processed_count, success_count, failure_count

    while not task_queue.empty():
        image_path = task_queue.get()
        try:
            print(f"[🧠 Thread-{threading.get_ident()}] 正在处理: {Path(image_path).name}")

            description = describe_image(image_path, client)

            with lock:
                processed_count += 1
                row = ws.max_row + 1

                # 获取图片原始尺寸
                with PILImage.open(image_path) as im:
                    orig_width, orig_height = im.size

                # 强制设置宽度为 200，计算新的高度
                thumb_width = 200
                scale = thumb_width / orig_width
                thumb_height = int(orig_height * scale)

                # 插入缩略图
                img = ExcelImage(image_path)
                img.width = thumb_width
                img.height = thumb_height
                ws.add_image(img, f"A{row}")

                # 自动设置行高和列宽
                ws.row_dimensions[row].height = thumb_height * 0.75
                ws.column_dimensions['A'].width = thumb_width / 7.5

                # 写入路径和描述
                ws.cell(row=row, column=2, value=image_path)

                if description:
                    ws.cell(row=row, column=3, value=description)
                    success_count += 1
                    print(f"[✅ Done] {Path(image_path).name} 描述成功 ✅（{processed_count}/{total_images}）")
                else:
                    ws.cell(row=row, column=3, value="API 调用失败")
                    failure_count += 1
                    print(f"[⚠️ Fail] {Path(image_path).name} 描述失败 ❌（{processed_count}/{total_images}）")

                progress_bar.update(1)

        finally:
            task_queue.task_done()


# ====== 加载图片 ======
print("🔍 正在扫描图片文件夹...")
for image_file in os.listdir(IMAGE_FOLDER):
    image_path = os.path.abspath(os.path.join(IMAGE_FOLDER, image_file))
    if Path(image_path).suffix.lower() in VALID_EXTENSIONS:
        task_queue.put(image_path)

total_images = task_queue.qsize()
print(f"🖼️ 共发现 {total_images} 张图片，准备启动 {MAX_THREADS} 个线程进行处理...\n")

# ====== 启动线程池并显示进度条 ======
threads = []
with tqdm(total=total_images, desc="📦 处理进度", ncols=100) as progress_bar:
    for _ in range(min(MAX_THREADS, total_images)):
        t = threading.Thread(target=worker, args=(progress_bar,))
        t.start()
        threads.append(t)

    for t in threads:
        t.join()

# ====== 保存 Excel 文件 ======
wb.save(OUTPUT_EXCEL)
end_time = time.time()
duration = round(end_time - start_time, 10)

# ====== 输出总结 ======
print("\n📊 处理完成总结：")
print(f"✅ 成功描述图片数量: {success_count}")
print(f"❌ 失败描述图片数量: {failure_count}")
print(f"⏱️ 总耗时: {duration} 秒")
print(f"📁 Excel 文件已保存至: {OUTPUT_EXCEL}")
print("🔄 处理完成，感谢使用！")